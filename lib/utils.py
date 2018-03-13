#!/usr/bin/env python
# -*- coding: utf-8 -*-

import IPy
import json
import time
import copy
import when
import arrow
import random
import logging
import openpyxl

from openpyxl.styles import Alignment
from tornado import web, gen, httpclient
from tornado.httputil import url_concat
from settings import COMMON_URL, STATE, BILL_PAY_STATE, PAY_TYPE, PAY_MD, ORDER_STATE


httpclient.AsyncHTTPClient.configure('tornado.simple_httpclient.SimpleAsyncHTTPClient', max_clients=300)


class APIError(web.HTTPError):
    '''
    自定义API异常
    '''
    def __init__(self, status_code=200, *args, **kwargs):
        super(APIError, self).__init__(status_code, *args, **kwargs)
        self.kwargs = kwargs

def dict_filter(target, attr=()):
    result = dict()
    for p in attr:
        if type(p) is dict:
            key = list(p.keys())[0]
            value = list(p.values())[0]
            result[value] = target[key] if key in target else ''
        elif p in target:
            result[p] = target[p]
    return result

def toNone(va):
    # empty to none
    return None if not va else va

def http_request(url, method='GET', **wargs):
    return httpclient.HTTPRequest(url=url, method=method, connect_timeout=10, request_timeout=10, **wargs)

def get_async_client():
    http_client = httpclient.AsyncHTTPClient()
    return http_client

async def fetch(http_client, request):
    r = await http_client.fetch(request)
    logging.info('\treq_url=%s\trequest_time=%s' % (r.effective_url, r.request_time))
    return r

async def async_common_api(path, params={}):
    url = url_concat(COMMON_URL + path, params)
    http_client = get_async_client()
    try:
        request = http_request(url)
        response = await fetch(http_client, request)
        response = json.loads(response.body.decode())
        return response
    except Exception as e:
        logging.error('url=%s, error=%s' % (url, e))
        raise APIError(errcode=10001, errmsg='公共接口请求失败')

def common_api(path, params={}):
    url = url_concat(COMMON_URL + path, params)
    http_client = httpclient.HTTPClient()
    try:
        request = http_request(url)
        response = http_client.fetch(request)
        response = json.loads(response.body.decode())
        return response
    except Exception as e:
        logging.error('url=%s, error=%s' % (url, e))
        raise APIError(errcode=10001, errmsg='公共接口请求失败')
    finally:
        http_client.close()

def common_post_api(path, params={}, method='POST'):
    url = COMMON_URL + path
    http_client = httpclient.HTTPClient()
    try:
        request = http_request(url, method=method, body=json.dumps(params))
        response = http_client.fetch(request)
        response = json.loads(response.body.decode())
        return response
    except Exception as e:
        logging.error('url=%s, error=%s' % (url, e))
        raise APIError(errcode=10001, errmsg='公共接口请求失败')
    finally:
        http_client.close()

def seconds_to_midnight():
    now = time.localtime()
    drawn = time.mktime(now[:3] + (0, 0, 0) + now[6:])
    return int(drawn + 86400 - time.time())

def is_valid_date(date_str):
    if not date_str:
        return True
    try:
        time.strptime(date_str, '%Y-%m-%d')
        return True
    except:
        return False

def is_valid_time(time_str):
    if not time_str:
        return True
    try:
        time.strptime(time_str, '%H:%M')
        return True
    except:
        return False

def is_valid_day_value(day_str):
    if not day_str:
        return True
    try:
        days = day_str.split(',')
        for day in days:
            if int(day) > 7 or int(day) < 1:
                return False
    except:
        return False
    else:
        return True

def is_valid_rt_ids_value(rt_ids_str):
    if not rt_ids_str:
        return True
    try:
        rt_ids = rt_ids_str.split(',')
        for rt_id in rt_ids:
            int(rt_id)
    except:
        return False
    else:
        return True

def is_ip_address(address):
    if not address:
        return True
    try:
        IPy.IP(address)
        return True
    except:
        return False

def get_day_of_week(day=None):
    '''
    取值1-7，对应周一至周日
    '''
    day = day if day else when.today()
    return day.weekday() + 1

def is_valid_pack(pack, today=None):
    today = today if today else when.today()
    valid_days = [int(day) for day in pack['day'].split(',')]
    day = get_day_of_week(today)

    if pack['state'] == STATE['valid'] and pack['start_date'] <= today <= pack['end_date'] and day in valid_days:
        return True

    return False

def future_time_by_hour(st, hour, format='HH:mm'):
    st = arrow.get(st, format)
    ed = st.replace(hours=hour)

    return ed.format(format)

def future_time_by_minute(st, minute, format='HH:mm'):
    st = arrow.get(st, format)
    ed = st.replace(minutes=minute)

    return ed.format(format)

def minute_distance(time_str_1, time_str_2):
    return (arrow.get(time_str_2, 'HH:mm') - arrow.get(time_str_1, 'HH:mm')).seconds / 60

def is_valid_fees(fees, store_st, store_ed):
    '''
    判断是否是有效的计费时段设置
    fees：某天全部计费时段
    store_st：商户营业开始时间
    store_ed：商户营业结束时间
    '''
    cfees = copy.deepcopy(fees)
    if not cfees:
        return False

    for f in cfees:
        if f['st'] < store_st:
            h, m = f['st'].split(':')
            f['st'] = '%s:%s'%(24+int(h), m)
        if f['ed'] <= store_st:
            h, m = f['ed'].split(':')
            f['ed'] = '%s:%s'%(24+int(h), m)

    if store_ed <= store_st:
        h, m = store_ed.split(':')
        store_ed = '%s:%s'%(24+int(h), m)

    cfees = sorted(cfees, key=lambda f: f['st'])

    start_st = cfees[0]['st']
    end_ed = cfees[len(cfees) - 1]['ed']

    if start_st != store_st or end_ed != store_ed:
        return False

    for index, fee in enumerate(cfees[1:]):
        st = fee['st']
        last_ed = cfees[index]['ed']
        if st != last_ed:
            return False

    return True

def is_hour_in_range(hour, start, end, tp='st'):
    # if start == end:
    #     raise APIError(errcode=50001, errmsg='计费方式设置有误')
    if start < end:
        if tp == 'st':
            if start <= hour < end:
                return True
        else:
            if start < hour <= end:
                return True

        return False

    if tp == 'st':
        if start <= hour < '24:00' or '00:00' <= hour < end:
            return True
    else:
        if start < hour < '24:00' or '00:00' <= hour <= end:
            return True

    return False

def get_time_bills(fees, st, ed):
    st = st[-5:]
    ed = ed[-5:]
    st_index = -1
    ed_index = -1

    for index, fee in enumerate(fees):
        if is_hour_in_range(st, fee['st'], fee['ed']):
            st_index = index
        if is_hour_in_range(ed, fee['st'], fee['ed'], 'ed'):
            ed_index = index

    if st_index == -1 or ed_index == -1:
        raise APIError(errcode=50001, errmsg='开台时间或关台时间设置有误')

    bills = []

    if st_index == ed_index:
        fee = fees[st_index]

        fee_id = fee['id']
        minute = minute_distance(st, ed)
        money_minute = int(fee['fee'] / 60)
        money = int(minute * fee['fee'] / 60)

        bills.append({
            'st': st,
            'ed': ed,
            'fee_id': fee_id,
            'minute': minute,
            'money_minute': money_minute,
            'money': money
        })

    elif ed_index > st_index:

        for index in range(st_index, ed_index + 1):
            fee = fees[index]
            fee_id = fee['id']

            if index == st_index:
                gen_st = st
                gen_ed = fee['ed']
            elif index == ed_index:
                gen_st = fee['st']
                gen_ed = ed
            else:
                gen_st = fee['st']
                gen_ed = fee['ed']

            minute = minute_distance(gen_st, gen_ed)
            money_minute = int(fee['fee'] / 60)
            money = int(minute * fee['fee'] / 60)

            bills.append({
                'st': gen_st,
                'ed': gen_ed,
                'fee_id': fee_id,
                'minute': minute,
                'money_minute': money_minute,
                'money': money
            })

    return bills

def gen_order_no():
    return when.now().strftime('%Y%m%d%H%M%S') + str(random.randint(0, 99)).zfill(2)

def gen_bill_no(prefix):
    return prefix + gen_order_no()

def get_pay_state(pay_type):
    pay_state = BILL_PAY_STATE['pay'] if pay_type == PAY_TYPE['current'] else BILL_PAY_STATE['unpay']
    return pay_state

def get_bill_pay_state(pay_md):
    '''支付方式如果是现金, 则是已支付'''
    if pay_md in (PAY_MD['cash'], PAY_MD['pos']):
        return BILL_PAY_STATE['pay']
    return BILL_PAY_STATE['unpay']

def get_today_start_end_time(st, ed):
    today = when.today()
    format_str = '%s %s'

    if st < ed:
        st_time_str = format_str % (today, st)
        ed_time_str = format_str % (today, ed)

        return st_time_str, ed_time_str

    tomorrow = when.tomorrow()
    st_time_str = format_str % (today, st)
    ed_time_str = format_str % (tomorrow, ed)

    return st_time_str, ed_time_str

def get_yesterday_start_end_time(st, ed):
    yesterday = when.yesterday()
    format_str = '%s %s'

    if st < ed:
        st_time_str = format_str % (yesterday, st)
        ed_time_str = format_str % (yesterday, ed)

        return st_time_str, ed_time_str

    today = when.today()
    st_time_str = format_str % (yesterday, st)
    ed_time_str = format_str % (today, ed)

    return st_time_str, ed_time_str

def get_7days_start_end_time(st, ed):
    today = when.today()
    seven_days_ago = when.past(days=7).date()
    format_str = '%s %s'

    if st < ed:
        st_time_str = format_str % (seven_days_ago, st)
        ed_time_str = format_str % (today, ed)

        return st_time_str, ed_time_str

    tomorrow = when.tomorrow()
    st_time_str = format_str % (seven_days_ago, st)
    ed_time_str = format_str % (tomorrow, ed)

    return st_time_str, ed_time_str

def get_30days_start_end_time(st, ed):
    today = when.today()
    thirty_days_ago = when.past(days=30).date()
    format_str = '%s %s'

    if st < ed:
        st_time_str = format_str % (thirty_days_ago, st)
        ed_time_str = format_str % (today, ed)

        return st_time_str, ed_time_str

    tomorrow = when.tomorrow()
    st_time_str = format_str % (thirty_days_ago, st)
    ed_time_str = format_str % (tomorrow, ed)

    return st_time_str, ed_time_str

def get_range_start_end_time(start_day, end_day, st, ed):
    format_str = '%s %s'
    start_day = start_day.date()

    if st < ed:
        st_time_str = format_str % (start_day, st)
        ed_time_str = format_str % (end_day.date(), ed)

        return st_time_str, ed_time_str

    end_day = end_day.replace(days=1)
    st_time_str = format_str % (start_day, st)
    ed_time_str = format_str % (end_day.date(), ed)

    return st_time_str, ed_time_str

def check_if_success_bill(pay_type, pay_md):
    return pay_type == PAY_TYPE['poster'] or pay_md in (PAY_MD['cash'], PAY_MD['pos'])

def export_xlsx(data, export_filename):
    '''
    data = [
        {'sheetname': 'sheet1', titles: ['title_1', 'title_2', 'title_3'], data: [[1,2,3], [2,3,4], [3,4,5]]},
        {'sheetname': 'sheet2', titles: ['title_1', 'title_1', 'title_1'], data: [[1,2,3], [2,3,4], [3,4,5]]}
    ]
    data是数组，如果长度大于1，则有多少个sheet
    '''
    assert isinstance(data, list)
    wb = openpyxl.Workbook()
    alignment = Alignment(wrap_text=True)
    ws_num = len(data)
    wss = []
    max_len = []   # 是数组的数组: [[], [], []], 记录每一个sheet, 每一列的最大长度, 导出时, 显示更正常
    for i in range(ws_num):
        if i == 0:
            ws = wb.active  # 第一个sheet是这么取的, 如果直接create_sheet, 生成的第一个sheet是空的
            ws.title = data[i].get('sheetname', '')
        else:
            ws = wb.create_sheet(data[i].get('sheetname', ''))
        for idx, title in enumerate(data[i].get('titles', [])):
            col = ord('A') + idx
            ws['%s1' % chr(col)] = title     # 写入标题
        # 初始化每一个sheet的每一列最大宽度为这个列名(汉字)的长度
        max_len.append([len(bytes(str(title), 'GBK')) for title in data[i].get('titles', [])])
        wss.append(ws)
    for idx, ws in enumerate(wss):
        data_lines = data[idx].get('data', [])
        for line, data_line in enumerate(data_lines):
            for col, data_col in enumerate(data_line):
                cur_col = ord('A') + col
                data_col = str(data_col)
                ws['%s%s' % (chr(cur_col), line+2)] = data_col
                ws['%s%s' % (chr(cur_col), line+2)].alignment = alignment
                if len(bytes(data_col, 'GBK')) > max_len[idx][col]:
                    max_len[idx][col] = len(bytes(data_col, 'GBK'))
        for colidx in range(len(data[idx].get('titles', []))):
            cur_col = ord('A') + colidx
            ws.column_dimensions['%s' % chr(cur_col)].width = max_len[idx][colidx]
    wb.save('%s' % export_filename)

def is_success_pay(tp, res):

    if int(res['errcode']) != 200:
        return False

    if ((tp=='ali' and
            (res.get('trade_status') == 'TRADE_SUCCESS') or
            int(res.get('code', 0))==10000) or
        (tp == 'wx' and
            res.get('return_code') == 'SUCCESS' and
            res.get('result_code') == 'SUCCESS' and
            res.get('trade_state') == 'SUCCESS')):
        return True

    return False
